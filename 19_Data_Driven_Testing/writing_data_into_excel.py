"""
openpyxl Methods-

    Method / Function	                Description

    load_workbook(filename)	            Loads an existing Excel workbook (.xlsx).
    Workbook()	                        Creates a new Excel workbook.
    wb.active	                        Returns the currently active worksheet.
    wb.sheetnames	                    Lists all sheet names in the workbook.
    wb["SheetName"]	                    Accesses a worksheet by its name.
    wb.create_sheet(title)	            Creates a new worksheet with the given title.
    del wb["SheetName"]	                Deletes a worksheet from the workbook.
    sheet["A1"]	                        Accesses a cell using Excel-style notation.
    sheet.cell(row, column)	            Accesses a cell using row and column indices.
    sheet["A1"].value	                Gets the value of a cell.
    sheet["A1"].font	                Gets or sets the font style of a cell.
    sheet.iter_rows()	                Iterates over rows in the worksheet.
    sheet.iter_cols()	                Iterates over columns in the worksheet.
    sheet.append([values])	            Appends a row of values to the worksheet.
    sheet.delete_rows(idx, amount)	    Deletes one or more rows starting at index idx.
    sheet.delete_cols(idx, amount)	    Deletes one or more columns starting at index idx.
    wb.save(filename)	                Saves the workbook to the specified filename.
    sheet.title	                        Gets or sets the title of a worksheet.
    sheet.max_row	                    Returns the maximum number of rows with data.
    sheet.max_column	                Returns the maximum number of columns with data.

"""

import openpyxl

# Same data in all the cells
file_1 = "E:/Jyotsna/Selenium_Learning/Selenium_Learning/19_Data_Driven_Testing/test_data.xlsx"
workbook_1 = openpyxl.load_workbook(file_1)
sheet_1 = workbook_1.active # or sheet = workbook["Sheet1"] ---> get active sheet from excel sheet

for r in range(1,6):
    for c in range(1,4):
        sheet_1.cell(r,c).value = "welcome"  # Writing the data from cell

workbook_1.save(file_1) #Save the file after entering the values

# Multiple data
file_2 = "E:/Jyotsna/Selenium_Learning/Selenium_Learning/19_Data_Driven_Testing/test_data_1.xlsx"
workbook_2 = openpyxl.load_workbook(file_2)
sheet_2 = workbook_2.active # or sheet = workbook["Sheet1"] ---> get active sheet from excel sheet

sheet_2.cell(1,1).value = "Emp_id"
sheet_2.cell(1,2).value = "name"
sheet_2.cell(1,3).value = "age"

sheet_2.cell(2,1).value = 187
sheet_2.cell(2,2).value = "john"
sheet_2.cell(2,3).value = 25

sheet_2.cell(3,1).value = 234
sheet_2.cell(3,2).value = "david"
sheet_2.cell(3,3).value = 28

sheet_2.cell(4,1).value = 123
sheet_2.cell(4,2).value = "john"
sheet_2.cell(4,3).value = 30

workbook_2.save(file_2)

