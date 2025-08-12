"""
A utility file in Selenium is a
    reusable helper module that contains
    common functions to simplify and streamline your test automation code.
    Instead of repeating logic across multiple test scripts, you centralize it in a utility class or module.

Common Functions in Selenium Utility Files-

1. Browser Setup & Teardown
    def init_driver():
        driver = webdriver.Chrome()
        driver.maximize_window()
        return driver

    def close_driver(driver):
        driver.quit()

2. Wait Helpers
    def wait_for_element(driver, by, locator, timeout=10):
        return WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, locator))
        )

    def wait_and_click(driver, by, locator, timeout=10):
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((by, locator))
        )
        element.click()

3. Element Interactions
    def enter_text(element, text):
        element.clear()
        element.send_keys(text)

    def get_text(element):
        return element.text

    def is_element_visible(driver, by, locator):
        try:
            return driver.find_element(by, locator).is_displayed()
        except:
            return False

4. Dropdown Handling
    from selenium.webdriver.support.ui import Select

    def select_by_visible_text(element, text):
        Select(element).select_by_visible_text(text)

    def select_by_value(element, value):
        Select(element).select_by_value(value)

5. Screenshot Capture
    def capture_screenshot(driver, name="screenshot"):
        driver.save_screenshot(f"screenshots/{name}.png")

6. Excel File Handling
    (If you're doing data-driven testing)

    def read_excel(file_path):
        # Returns list of dictionaries from Excel rows
        ...

    def write_excel(file_path, results):
        # Writes results to Excel
        ...

7. Alert Handling
    def accept_alert(driver):
        WebDriverWait(driver, 5).until(EC.alert_is_present())
        driver.switch_to.alert.accept()

    def dismiss_alert(driver):
        WebDriverWait(driver, 5).until(EC.alert_is_present())
        driver.switch_to.alert.dismiss()

8. Frame & Window Switching
    def switch_to_frame(driver, frame_reference):
        driver.switch_to.frame(frame_reference)

    def switch_to_window(driver, index=0):
        driver.switch_to.window(driver.window_handles[index])

9. Scrolling
    def scroll_to_element(driver, element):
        driver.execute_script("arguments[0].scrollIntoView();", element)

    def scroll_by_offset(driver, x=0, y=500):
        driver.execute_script(f"window.scrollBy({x}, {y});")

10. Logging (Optional)
    def log(message):
        print(f"[LOG] {message}")

Example of a utility file is given below
"""

import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(file,sheetName,row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row_num,column_num).value

def writeData(file,sheetName,row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row_num,column_num).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green_fill = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type='solid')
    sheet.cell(row_num, column_num).fill = green_fill
    workbook.save(file)


def fillRedColor(file, sheetName, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red_fill = PatternFill(start_color='ff0000',
                             end_color='ff0000',
                             fill_type='solid')
    sheet.cell(row_num, column_num).fill = red_fill
    workbook.save(file)
