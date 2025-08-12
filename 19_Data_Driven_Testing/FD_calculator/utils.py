import openpyxl

def read_test_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            'Principal': row[0],
            'Rate': row[1],
            'Tenure': row[2],
            'Expected': row[3]
        })
    return data

def write_result(file_path, results):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    for i, result in enumerate(results, start=2):
        sheet.cell(row=i, column=5).value = result  # Assuming column E is for a result
    wb.save(file_path)
