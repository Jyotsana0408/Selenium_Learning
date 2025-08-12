"""

Data inside data.xlsx

Bookname	PurchasedDate	Amount	Location
Selenium	11-Aug-25	    350	    Africa
Java	    11-Aug-25	    200	    Africa
Python	    11-Aug-25	    250	    Asia
Jmeter	    11-Aug-25	    150	    Asia
C	        11-Aug-25	    300	    Asia


"""
import openpyxl

# File-->Workbook-->Sheets-->Rows-->Cells
file = "E:/Jyotsna/Selenium_Learning/Selenium_Learning/19_Data_Driven_Testing/data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row # count the number of rows in Excelsheet
columns = sheet.max_column # count the number of columns in Excelsheet

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end="     ") # Read the data from cell
    print()

